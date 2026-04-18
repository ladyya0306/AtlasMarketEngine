#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build deterministic supply snapshot catalogs for research.
"""

from __future__ import annotations

import argparse
import copy
import csv
import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.line_b_library_builder import build_governance_snapshot

DEFAULT_CATALOG_CFG = ROOT / "config" / "supply_snapshot_catalog_v1.yaml"
DEFAULT_RESULTS_ROOT = ROOT / "results" / "supply_snapshot_catalog"

SNAPSHOT_FIELDS = [
    "snapshot_id",
    "catalog_id",
    "catalog_version",
    "recommended_snapshot_id",
    "profile_pack_artifact",
    "structure_family",
    "size_profile",
    "recommended_use",
    "experiment_mode",
    "snapshot_status",
    "status_summary",
    "status_failures",
    "status_warnings",
    "total_selected_supply",
    "supply_bucket_count",
    "total_agents_profiled",
    "demand_bucket_count",
    "graph_ok",
    "budget_ok",
    "reverse_coverage_ok",
    "runtime_parent_ok",
    "zero_affordable_bucket_count",
    "thin_affordable_bucket_count",
    "competition_hotspot_bucket_count",
    "missing_primary_supply_bucket_count",
    "missing_primary_demand_bucket_count",
    "unaffordable_primary_edge_count",
]

SUPPLY_BUCKET_FIELDS = [
    "snapshot_id",
    "bucket_id",
    "tier",
    "zone",
    "is_school_district",
    "property_type_bucket",
    "product_segment",
    "household_fit",
    "value_logic",
    "transaction_liquidity_band",
    "reverse_primary_required",
    "price_low",
    "price_high",
    "bedroom_low",
    "bedroom_high",
    "building_area_low",
    "building_area_high",
    "quality_low",
    "quality_high",
    "count_abundant_base",
    "count_scarce_base",
    "count_selected_mode",
]

SUPPLY_PROPERTY_FIELDS = [
    "snapshot_id",
    "property_id",
    "bucket_id",
    "synthetic_order",
    "zone",
    "is_school_district",
    "property_type_bucket",
    "product_segment",
    "household_fit",
    "value_logic",
    "transaction_liquidity_band",
    "list_price",
    "building_area",
    "bedrooms",
    "quality_score",
]

DEMAND_COVERAGE_FIELDS = [
    "snapshot_id",
    "demand_bucket_id",
    "role_side",
    "buyer_count",
    "agent_type",
    "info_delay_months",
    "target_zone",
    "need_school_district",
    "property_type_target",
    "target_buy_price_low",
    "target_buy_price_high",
    "buyer_max_price_low",
    "buyer_max_price_high",
    "primary_supply_bucket_ids",
    "secondary_supply_bucket_ids",
    "eligible_supply_bucket_ids",
    "affordable_supply_bucket_ids",
    "primary_selected_supply_count",
    "eligible_selected_supply_count",
    "affordable_selected_supply_count",
    "competition_ratio",
    "coverage_status",
    "coverage_reason",
]


def _load_yaml(path: Path) -> Dict[str, Any]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError(f"YAML payload must be dict: {path}")
    return payload


def _resolve_profile_pack(path: Path) -> Dict[str, Any]:
    payload = _load_yaml(path)
    pack = payload.get("profiled_market_mode", payload)
    if not isinstance(pack, dict):
        raise ValueError(f"profile pack is invalid: {path}")
    return copy.deepcopy(pack)


def _safe_range(raw: Any, default_low: float = 0.0, default_high: float = 0.0) -> Tuple[float, float]:
    if isinstance(raw, (list, tuple)) and len(raw) >= 2:
        low = float(raw[0] or default_low)
        high = float(raw[1] or default_high)
        if low > high:
            low, high = high, low
        return low, high
    return float(default_low), float(default_high)


def _bucket_value(bucket: Dict[str, Any], key: str, default: Any = None) -> Any:
    if not isinstance(bucket, dict):
        return default
    if bucket.get(key) not in (None, ""):
        return bucket.get(key)
    market_profile = bucket.get("market_profile", {}) or {}
    if isinstance(market_profile, dict) and market_profile.get(key) not in (None, ""):
        return market_profile.get(key)
    return default


def _derive_bucket_tier(bucket_id: str, bucket: Dict[str, Any]) -> str:
    property_type_bucket = str(_bucket_value(bucket, "property_type_bucket", "") or "").strip().upper()
    price_low, price_high = _safe_range(_bucket_value(bucket, "price_range", []), 0.0, 0.0)
    price_mid = (price_low + price_high) / 2.0
    name = str(bucket_id or "").upper()
    if "LUXURY" in property_type_bucket or "LUXURY" in name or price_mid >= 10_000_000:
        return "luxury"
    if property_type_bucket == "JUST" or "STARTER" in name or price_mid <= 2_000_000:
        return "starter"
    return "improve"


def _scaled_bucket_count(base_count: int, family_scale: float, size_scale: float) -> int:
    if int(base_count or 0) <= 0:
        return 0
    scaled = int(round(float(base_count) * float(family_scale) * float(size_scale)))
    return max(1, scaled)


def _interpolate(low: float, high: float, ordinal: int, total: int) -> float:
    if total <= 1:
        return round((float(low) + float(high)) / 2.0, 2)
    ratio = (float(ordinal) - 0.5) / float(total)
    return round(float(low) + (float(high) - float(low)) * ratio, 2)


def _join_ids(values: Iterable[str]) -> str:
    output = [str(value).strip() for value in values if str(value or "").strip()]
    return "|".join(output)


def _normalize_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _bool_to_int(value: Any) -> int:
    return 1 if bool(value) else 0


def _expand_supply_rows(
    *,
    snapshot_id: str,
    supply_buckets: List[Dict[str, Any]],
    property_bucket_map: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    bucket_rows: List[Dict[str, Any]] = []
    property_rows: List[Dict[str, Any]] = []
    for bucket in sorted(supply_buckets, key=lambda item: str(item.get("bucket_id", ""))):
        bucket_id = str(bucket.get("bucket_id", "") or "").strip()
        source_bucket = property_bucket_map.get(bucket_id, {}) or {}
        market_profile = source_bucket.get("market_profile", {}) or {}
        property_profile = source_bucket.get("property_profile", {}) or {}
        price_low, price_high = _safe_range(bucket.get("price_range"), 0.0, 0.0)
        bedroom_low, bedroom_high = _safe_range(market_profile.get("bedroom_range"), 0.0, 0.0)
        area_low, area_high = _safe_range(market_profile.get("building_area_range"), 0.0, 0.0)
        quality_low, quality_high = _safe_range(market_profile.get("quality_range"), 0.0, 0.0)
        count_by_mode = source_bucket.get("count_by_supply_mode", {}) or {}
        count_selected = _normalize_int(bucket.get("count_selected_mode"))
        bucket_row = {
            "snapshot_id": snapshot_id,
            "bucket_id": bucket_id,
            "tier": _derive_bucket_tier(bucket_id, source_bucket),
            "zone": str(bucket.get("zone", "") or ""),
            "is_school_district": _bool_to_int(bucket.get("is_school_district")),
            "property_type_bucket": str(bucket.get("property_type_bucket", "") or ""),
            "product_segment": str(property_profile.get("product_segment", "") or ""),
            "household_fit": str(property_profile.get("household_fit", "") or ""),
            "value_logic": str(property_profile.get("value_logic", "") or ""),
            "transaction_liquidity_band": str(property_profile.get("transaction_liquidity_band", "") or ""),
            "reverse_primary_required": _bool_to_int(source_bucket.get("reverse_primary_required")),
            "price_low": price_low,
            "price_high": price_high,
            "bedroom_low": _normalize_int(bedroom_low),
            "bedroom_high": _normalize_int(bedroom_high),
            "building_area_low": area_low,
            "building_area_high": area_high,
            "quality_low": quality_low,
            "quality_high": quality_high,
            "count_abundant_base": _normalize_int(count_by_mode.get("abundant")),
            "count_scarce_base": _normalize_int(count_by_mode.get("scarce")),
            "count_selected_mode": count_selected,
        }
        bucket_rows.append(bucket_row)
        for ordinal in range(1, count_selected + 1):
            property_rows.append(
                {
                    "snapshot_id": snapshot_id,
                    "property_id": f"{snapshot_id}__{bucket_id}__{ordinal:03d}",
                    "bucket_id": bucket_id,
                    "synthetic_order": ordinal,
                    "zone": bucket_row["zone"],
                    "is_school_district": bucket_row["is_school_district"],
                    "property_type_bucket": bucket_row["property_type_bucket"],
                    "product_segment": bucket_row["product_segment"],
                    "household_fit": bucket_row["household_fit"],
                    "value_logic": bucket_row["value_logic"],
                    "transaction_liquidity_band": bucket_row["transaction_liquidity_band"],
                    "list_price": _interpolate(price_low, price_high, ordinal, count_selected),
                    "building_area": _interpolate(area_low, area_high, ordinal, count_selected)
                    if area_high > 0
                    else 0.0,
                    "bedrooms": _normalize_int(_interpolate(bedroom_low, bedroom_high, ordinal, count_selected))
                    if bedroom_high > 0
                    else 0,
                    "quality_score": _interpolate(quality_low, quality_high, ordinal, count_selected)
                    if quality_high > 0
                    else 0.0,
                }
            )
    return bucket_rows, property_rows


def _build_demand_coverage_rows(
    *,
    snapshot_id: str,
    snapshot: Dict[str, Any],
    profile_pack: Dict[str, Any],
) -> List[Dict[str, Any]]:
    demand_buckets = snapshot.get("demand_library", {}).get("buckets", []) or []
    supply_buckets = snapshot.get("supply_library", {}).get("buckets", []) or []
    budget_rows = snapshot.get("budget_consistency_report", {}).get("rows", []) or []
    competition_rows = snapshot.get("competition_control_report", {}).get(
        snapshot.get("supply_library", {}).get("experiment_mode", "abundant"),
        [],
    ) or []
    graph = profile_pack.get("compatibility_graph_v1", {}) or {}
    demand_edges = graph.get("demand_to_supply", []) or []
    primary_map: Dict[str, List[str]] = {}
    secondary_map: Dict[str, List[str]] = {}
    for edge in demand_edges:
        demand_bucket_id = str(edge.get("demand_bucket_id", "") or "").strip()
        supply_bucket_id = str(edge.get("supply_bucket_id", "") or "").strip()
        relation_type = str(edge.get("relation_type", "primary") or "primary").strip().lower()
        if not demand_bucket_id or not supply_bucket_id:
            continue
        target_map = primary_map if relation_type == "primary" else secondary_map
        target_map.setdefault(demand_bucket_id, []).append(supply_bucket_id)
    supply_count_map = {
        str(bucket.get("bucket_id", "") or "").strip(): _normalize_int(bucket.get("count_selected_mode"))
        for bucket in supply_buckets
    }
    budget_map = {
        str(row.get("agent_bucket_id", "") or "").strip(): row
        for row in budget_rows
    }
    competition_map = {
        str(row.get("agent_bucket_id", "") or "").strip(): row
        for row in competition_rows
    }
    output: List[Dict[str, Any]] = []
    for bucket in demand_buckets:
        demand_bucket_id = str(bucket.get("bucket_id", "") or "").strip()
        budget_row = budget_map.get(demand_bucket_id, {}) or {}
        eligible_supply = budget_row.get("eligible_supply", []) or []
        eligible_supply_ids = [str(item.get("property_bucket_id", "") or "").strip() for item in eligible_supply]
        affordable_supply_ids = [
            str(item.get("property_bucket_id", "") or "").strip()
            for item in eligible_supply
            if bool(item.get("affordable_any_supply"))
        ]
        primary_ids = primary_map.get(demand_bucket_id, [])
        secondary_ids = secondary_map.get(demand_bucket_id, [])
        affordable_selected_supply_count = sum(supply_count_map.get(bucket_id, 0) for bucket_id in affordable_supply_ids)
        eligible_selected_supply_count = sum(supply_count_map.get(bucket_id, 0) for bucket_id in eligible_supply_ids)
        primary_selected_supply_count = sum(supply_count_map.get(bucket_id, 0) for bucket_id in primary_ids)
        coverage_status = "ok"
        coverage_reason = ""
        if affordable_selected_supply_count <= 0:
            coverage_status = "zero_affordable"
            coverage_reason = "没有买得起的可匹配供应"
        elif affordable_selected_supply_count < 2:
            coverage_status = "thin_affordable"
            coverage_reason = "买得起的可匹配供应过薄"
        output.append(
            {
                "snapshot_id": snapshot_id,
                "demand_bucket_id": demand_bucket_id,
                "role_side": str(bucket.get("role_side", "") or ""),
                "buyer_count": _normalize_int(bucket.get("count")),
                "agent_type": str(bucket.get("agent_type", "") or ""),
                "info_delay_months": _normalize_int(bucket.get("info_delay_months")),
                "target_zone": str(bucket.get("target_zone", "") or ""),
                "need_school_district": _bool_to_int(bucket.get("need_school_district")),
                "property_type_target": str(bucket.get("property_type_target", "") or ""),
                "target_buy_price_low": _safe_range(bucket.get("target_buy_price_range"), 0.0, 0.0)[0],
                "target_buy_price_high": _safe_range(bucket.get("target_buy_price_range"), 0.0, 0.0)[1],
                "buyer_max_price_low": _safe_range(bucket.get("max_price_range"), 0.0, 0.0)[0],
                "buyer_max_price_high": _safe_range(bucket.get("max_price_range"), 0.0, 0.0)[1],
                "primary_supply_bucket_ids": _join_ids(primary_ids),
                "secondary_supply_bucket_ids": _join_ids(secondary_ids),
                "eligible_supply_bucket_ids": _join_ids(eligible_supply_ids),
                "affordable_supply_bucket_ids": _join_ids(affordable_supply_ids),
                "primary_selected_supply_count": primary_selected_supply_count,
                "eligible_selected_supply_count": eligible_selected_supply_count,
                "affordable_selected_supply_count": affordable_selected_supply_count,
                "competition_ratio": round(
                    float((competition_map.get(demand_bucket_id, {}) or {}).get("buyer_to_supply_ratio", 0.0) or 0.0),
                    4,
                ),
                "coverage_status": coverage_status,
                "coverage_reason": coverage_reason,
            }
        )
    return output


def _snapshot_status(snapshot: Dict[str, Any], demand_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    failures: List[str] = []
    warnings: List[str] = []
    budget_ok = bool(snapshot.get("budget_consistency_report", {}).get("ok"))
    graph_ok = bool(snapshot.get("graph_consistency_report", {}).get("ok"))
    reverse_coverage_ok = bool(snapshot.get("reverse_coverage_report", {}).get("ok"))
    runtime_parent_ok = bool(snapshot.get("runtime_parent_bucket_report", {}).get("ok"))
    graph_report = snapshot.get("graph_consistency_report", {}) or {}
    reverse_report = snapshot.get("reverse_coverage_report", {}) or {}
    if not budget_ok:
        failures.append("预算覆盖存在断裂")
    if not graph_ok:
        failures.append("供需画像图存在主链缺口")
    if not reverse_coverage_ok:
        failures.append("部分供应桶没有需求侧主覆盖")
    if not runtime_parent_ok:
        failures.append("运行期父桶映射不完整")
    zero_affordable = [row["demand_bucket_id"] for row in demand_rows if row.get("coverage_status") == "zero_affordable"]
    thin_affordable = [row["demand_bucket_id"] for row in demand_rows if row.get("coverage_status") == "thin_affordable"]
    competition_hotspots = [
        row["demand_bucket_id"]
        for row in demand_rows
        if float(row.get("competition_ratio", 0.0) or 0.0) >= 1.0
    ]
    if zero_affordable:
        failures.append(f"买得起的对位供应为 0：{_join_ids(zero_affordable)}")
    if thin_affordable:
        warnings.append(f"买得起的对位供应偏薄：{_join_ids(thin_affordable)}")
    if competition_hotspots:
        warnings.append(f"高竞争热点需求桶：{_join_ids(competition_hotspots)}")
    status = "pass"
    if failures:
        status = "fail"
    elif warnings:
        status = "warn"
    return {
        "snapshot_status": status,
        "status_summary": failures[0] if failures else (warnings[0] if warnings else "结构覆盖通过"),
        "status_failures": failures,
        "status_warnings": warnings,
        "budget_ok": budget_ok,
        "graph_ok": graph_ok,
        "reverse_coverage_ok": reverse_coverage_ok,
        "runtime_parent_ok": runtime_parent_ok,
        "zero_affordable_bucket_count": len(zero_affordable),
        "thin_affordable_bucket_count": len(thin_affordable),
        "competition_hotspot_bucket_count": len(competition_hotspots),
        "missing_primary_supply_bucket_count": len(graph_report.get("missing_primary_supply_buckets", []) or []),
        "missing_primary_demand_bucket_count": len(graph_report.get("missing_primary_demand_buckets", []) or []),
        "unaffordable_primary_edge_count": len(graph_report.get("unaffordable_primary_edges", []) or []),
        "uncovered_primary_supply_bucket_count": len(
            reverse_report.get("uncovered_primary_supply_buckets", []) or []
        ),
    }


def _coerce_cell(value: Any) -> Any:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _infer_sqlite_type(rows: List[Dict[str, Any]], field: str) -> str:
    values = [row.get(field) for row in rows if row.get(field) not in (None, "")]
    if not values:
        return "TEXT"
    if all(isinstance(value, bool) for value in values):
        return "INTEGER"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in values):
        return "INTEGER"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
        return "REAL" if any(isinstance(value, float) for value in values) else "INTEGER"
    return "TEXT"


def _write_sqlite(
    *,
    db_path: Path,
    tables: Dict[str, Dict[str, Any]],
) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        for table_name, payload in tables.items():
            fields = payload["fields"]
            rows = payload["rows"]
            columns = ", ".join(
                f'"{field}" {_infer_sqlite_type(rows, field)}' for field in fields
            )
            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            conn.execute(f'CREATE TABLE "{table_name}" ({columns})')
            if not rows:
                continue
            placeholders = ", ".join("?" for _ in fields)
            quoted_fields = ", ".join(f'"{field}"' for field in fields)
            sql = f'INSERT INTO "{table_name}" ({quoted_fields}) VALUES ({placeholders})'
            conn.executemany(
                sql,
                [[_coerce_cell(row.get(field)) for field in fields] for row in rows],
            )
        conn.commit()


def _write_csv(*, path: Path, fieldnames: List[str], rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _coerce_cell(row.get(field)) for field in fieldnames})


def _append_sheet(
    *,
    workbook: Workbook,
    title: str,
    fieldnames: List[str],
    rows: List[Dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, field in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=column_index, value=field)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fieldnames, start=1):
            sheet.cell(row=row_index, column=column_index, value=_coerce_cell(row.get(field)))
    sheet.freeze_panes = "A2"
    for column_index, field in enumerate(fieldnames, start=1):
        max_width = max(
            len(str(field)),
            *[len(str(_coerce_cell(row.get(field)) or "")) for row in rows[:200]],
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(12, max_width + 2),
            48,
        )


def _write_xlsx(
    *,
    xlsx_path: Path,
    snapshots: List[Dict[str, Any]],
    supply_buckets: List[Dict[str, Any]],
    supply_properties: List[Dict[str, Any]],
    demand_coverage: List[Dict[str, Any]],
) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_sheet(workbook=workbook, title="snapshots", fieldnames=SNAPSHOT_FIELDS, rows=snapshots)
    _append_sheet(workbook=workbook, title="supply_buckets", fieldnames=SUPPLY_BUCKET_FIELDS, rows=supply_buckets)
    _append_sheet(
        workbook=workbook,
        title="supply_properties",
        fieldnames=SUPPLY_PROPERTY_FIELDS,
        rows=supply_properties,
    )
    _append_sheet(
        workbook=workbook,
        title="demand_coverage",
        fieldnames=DEMAND_COVERAGE_FIELDS,
        rows=demand_coverage,
    )
    workbook.save(xlsx_path)


def build_snapshot_catalog(
    *,
    catalog_config_path: Path,
    out_root: Path,
    seed: int,
) -> Path:
    catalog_cfg = _load_yaml(catalog_config_path)
    catalog_meta = catalog_cfg.get("catalog", {}) or {}
    base_cfg = catalog_cfg.get("base_profile_pack", {}) or {}
    structure_families = catalog_cfg.get("structure_families", {}) or {}
    size_profiles = catalog_cfg.get("size_profiles", {}) or {}
    presets = catalog_cfg.get("presets", []) or []
    bucket_tiers = catalog_cfg.get("property_bucket_tiers", {}) or {}
    experiment_mode = str(base_cfg.get("experiment_mode", "abundant") or "abundant")
    base_pack_path = (ROOT / str(base_cfg.get("path", "") or "")).resolve()
    base_profile_pack = _resolve_profile_pack(base_pack_path)
    snapshots: List[Dict[str, Any]] = []
    supply_bucket_rows: List[Dict[str, Any]] = []
    supply_property_rows: List[Dict[str, Any]] = []
    demand_coverage_rows: List[Dict[str, Any]] = []
    snapshot_payloads: Dict[str, Dict[str, Any]] = {}
    exported_profile_packs: Dict[str, Dict[str, Any]] = {}
    severity_rank = {"pass": 0, "warn": 1, "fail": 2}
    overall_status = "pass"
    for preset in presets:
        snapshot_id = str(preset.get("snapshot_id", "") or "").strip()
        structure_family = str(preset.get("structure_family", "") or "").strip()
        size_profile = str(preset.get("size_profile", "") or "").strip()
        recommended_use = str(preset.get("recommended_use", "") or "")
        family_cfg = structure_families.get(structure_family, {}) or {}
        size_cfg = size_profiles.get(size_profile, {}) or {}
        size_scale = float(size_cfg.get("count_scale", 1.0) or 1.0)
        family_scale_map = family_cfg.get("tier_scale_map", {}) or {}
        profile_pack = copy.deepcopy(base_profile_pack)
        property_bucket_map = profile_pack.get("property_profile_buckets", {}) or {}
        if not isinstance(property_bucket_map, dict):
            raise ValueError("property_profile_buckets must be a dict")
        for bucket_id, bucket in property_bucket_map.items():
            if not isinstance(bucket, dict):
                continue
            tier = str(bucket_tiers.get(bucket_id) or _derive_bucket_tier(str(bucket_id), bucket))
            family_scale = float(family_scale_map.get(tier, 1.0) or 1.0)
            count_by_mode = copy.deepcopy(bucket.get("count_by_supply_mode", {}) or {})
            base_count = _normalize_int(count_by_mode.get(experiment_mode))
            count_by_mode[experiment_mode] = _scaled_bucket_count(base_count, family_scale, size_scale)
            bucket["count_by_supply_mode"] = count_by_mode
        exported_profile_packs[snapshot_id] = {"profiled_market_mode": copy.deepcopy(profile_pack)}
        snapshot = build_governance_snapshot(
            profile_pack=profile_pack,
            profile_pack_path=str(base_pack_path),
            experiment_mode=experiment_mode,
            seed=seed,
            group_id=snapshot_id,
            months=3,
            agent_count=12,
        )
        bucket_rows, property_rows = _expand_supply_rows(
            snapshot_id=snapshot_id,
            supply_buckets=snapshot.get("supply_library", {}).get("buckets", []) or [],
            property_bucket_map=property_bucket_map,
        )
        coverage_rows = _build_demand_coverage_rows(
            snapshot_id=snapshot_id,
            snapshot=snapshot,
            profile_pack=profile_pack,
        )
        status_info = _snapshot_status(snapshot, coverage_rows)
        snapshot_row = {
            "snapshot_id": snapshot_id,
            "catalog_id": str(catalog_meta.get("catalog_id", "") or ""),
            "catalog_version": _normalize_int(catalog_meta.get("version")),
            "recommended_snapshot_id": str(catalog_meta.get("recommended_snapshot_id", "") or ""),
            "profile_pack_artifact": f"profile_packs/{snapshot_id}.yaml",
            "structure_family": structure_family,
            "size_profile": size_profile,
            "recommended_use": recommended_use,
            "experiment_mode": experiment_mode,
            "snapshot_status": status_info["snapshot_status"],
            "status_summary": status_info["status_summary"],
            "status_failures": _join_ids(status_info["status_failures"]),
            "status_warnings": _join_ids(status_info["status_warnings"]),
            "total_selected_supply": _normalize_int(snapshot.get("supply_library", {}).get("total_selected_supply")),
            "supply_bucket_count": _normalize_int(snapshot.get("supply_library", {}).get("bucket_count")),
            "total_agents_profiled": _normalize_int(snapshot.get("demand_library", {}).get("total_agents_profiled")),
            "demand_bucket_count": _normalize_int(snapshot.get("demand_library", {}).get("bucket_count")),
            "graph_ok": _bool_to_int(status_info["graph_ok"]),
            "budget_ok": _bool_to_int(status_info["budget_ok"]),
            "reverse_coverage_ok": _bool_to_int(status_info["reverse_coverage_ok"]),
            "runtime_parent_ok": _bool_to_int(status_info["runtime_parent_ok"]),
            "zero_affordable_bucket_count": status_info["zero_affordable_bucket_count"],
            "thin_affordable_bucket_count": status_info["thin_affordable_bucket_count"],
            "competition_hotspot_bucket_count": status_info["competition_hotspot_bucket_count"],
            "missing_primary_supply_bucket_count": status_info["missing_primary_supply_bucket_count"],
            "missing_primary_demand_bucket_count": status_info["missing_primary_demand_bucket_count"],
            "unaffordable_primary_edge_count": status_info["unaffordable_primary_edge_count"],
        }
        snapshots.append(snapshot_row)
        supply_bucket_rows.extend(bucket_rows)
        supply_property_rows.extend(property_rows)
        demand_coverage_rows.extend(coverage_rows)
        snapshot_payloads[snapshot_id] = {
            "snapshot_row": snapshot_row,
            "status_info": status_info,
            "governance_snapshot": snapshot,
            "supply_buckets": bucket_rows,
            "demand_coverage": coverage_rows,
        }
        if severity_rank[status_info["snapshot_status"]] > severity_rank[overall_status]:
            overall_status = status_info["snapshot_status"]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = out_root / f"supply_snapshot_catalog_{timestamp}_{overall_status}"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "snapshots").mkdir(parents=True, exist_ok=True)
    (out_dir / "csv").mkdir(parents=True, exist_ok=True)
    (out_dir / "profile_packs").mkdir(parents=True, exist_ok=True)
    manifest = {
        "catalog": catalog_meta,
        "catalog_config_path": str(catalog_config_path),
        "base_profile_pack_path": str(base_pack_path),
        "experiment_mode": experiment_mode,
        "seed": seed,
        "overall_status": overall_status,
        "snapshot_count": len(snapshots),
        "output_dir": str(out_dir),
    }
    (out_dir / "catalog_config.snapshot.yaml").write_text(
        catalog_config_path.read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    (out_dir / "base_profile_pack.snapshot.yaml").write_text(
        base_pack_path.read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    for snapshot_id, profile_pack_payload in exported_profile_packs.items():
        (out_dir / "profile_packs" / f"{snapshot_id}.yaml").write_text(
            yaml.safe_dump(profile_pack_payload, allow_unicode=True, sort_keys=False),
            encoding="utf-8",
        )
    for snapshot_id, payload in snapshot_payloads.items():
        (out_dir / "snapshots" / f"{snapshot_id}.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_csv(path=out_dir / "csv" / "snapshots.csv", fieldnames=SNAPSHOT_FIELDS, rows=snapshots)
    _write_csv(
        path=out_dir / "csv" / "supply_buckets.csv",
        fieldnames=SUPPLY_BUCKET_FIELDS,
        rows=supply_bucket_rows,
    )
    _write_csv(
        path=out_dir / "csv" / "supply_properties.csv",
        fieldnames=SUPPLY_PROPERTY_FIELDS,
        rows=supply_property_rows,
    )
    _write_csv(
        path=out_dir / "csv" / "demand_coverage.csv",
        fieldnames=DEMAND_COVERAGE_FIELDS,
        rows=demand_coverage_rows,
    )
    _write_sqlite(
        db_path=out_dir / "supply_snapshot_catalog.db",
        tables={
            "snapshot_catalog": {"fields": SNAPSHOT_FIELDS, "rows": snapshots},
            "snapshot_supply_buckets": {"fields": SUPPLY_BUCKET_FIELDS, "rows": supply_bucket_rows},
            "snapshot_supply_properties": {
                "fields": SUPPLY_PROPERTY_FIELDS,
                "rows": supply_property_rows,
            },
            "snapshot_demand_coverage": {
                "fields": DEMAND_COVERAGE_FIELDS,
                "rows": demand_coverage_rows,
            },
        },
    )
    _write_xlsx(
        xlsx_path=out_dir / "supply_snapshot_catalog.xlsx",
        snapshots=snapshots,
        supply_buckets=supply_bucket_rows,
        supply_properties=supply_property_rows,
        demand_coverage=demand_coverage_rows,
    )
    return out_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="Build deterministic supply snapshot catalog artifacts.")
    parser.add_argument(
        "--catalog-config-path",
        type=Path,
        default=DEFAULT_CATALOG_CFG,
        help="Path to the supply snapshot catalog YAML.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=DEFAULT_RESULTS_ROOT,
        help="Directory under results/ where catalog artifacts will be written.",
    )
    parser.add_argument("--seed", type=int, default=606, help="Seed recorded into snapshot identity.")
    args = parser.parse_args()
    out_dir = build_snapshot_catalog(
        catalog_config_path=args.catalog_config_path.resolve(),
        out_root=args.out_dir.resolve(),
        seed=int(args.seed),
    )
    print(str(out_dir))


if __name__ == "__main__":
    main()
